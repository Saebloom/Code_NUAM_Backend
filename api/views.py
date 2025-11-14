# api/views.py
import logging
import csv
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib.auth import authenticate
from django.contrib.auth.models import Group
from django.core.cache import cache
from django.db import transaction
from django.db import models
from django.utils import timezone

from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework import viewsets, status, generics
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, IsAdminUser, IsAuthenticated
from rest_framework_simplejwt.tokens import RefreshToken

from .models import (
    Estado, Instrumento, Mercado, Archivo,
    Calificacion, CalificacionTributaria, FactorTributario,
    Log, Auditoria, Usuario, Respaldo
)

from .serializers import (
    UserSerializer, EstadoSerializer, InstrumentoSerializer,
    MercadoSerializer, ArchivoSerializer, CalificacionSerializer,
    CalificacionTributariaSerializer, FactorTributarioSerializer,
    LogSerializer, AuditoriaSerializer, 
    CurrentUserSerializer, RespaldoSerializer
)
from .permissions import IsAdminOrReadOnly, IsOwnerOrAdmin

logger = logging.getLogger('api')

# =============================================================
# VISTAS DE AUTENTICACIÓN
# =============================================================

@api_view(['POST'])
@permission_classes([AllowAny])
def login_nuam(request):
    """
    Login corporativo personalizado.
    """
    username = request.data.get('username', '').lower().strip()
    password = request.data.get('password')

    if not username.endswith('@nuam.cl'):
        return Response(
            {"detail": "Solo se permiten correos corporativos @nuam.cl"},
            status=status.HTTP_403_FORBIDDEN
        )
    
    user = authenticate(username=username, password=password) 
    
    if user is not None:
        if not user.is_active:
            return Response(
                {"detail": "La cuenta de usuario está deshabilitada."},
                status=status.HTTP_403_FORBIDDEN
            )
        
        refresh = RefreshToken.for_user(user)
        # Devolvemos el usuario (lo necesita index.html)
        serializer = CurrentUserSerializer(user)
        return Response({
            'refresh': str(refresh),
            'access': str(refresh.access_token),
            'user': serializer.data
        })
    else:
        return Response(
            {"detail": "Credenciales inválidas."},
            status=status.HTTP_401_UNAUTHORIZED
        )

# =============================================================
# VISTAS DE USUARIOS (ADMIN)
# =============================================================

class UserViewSet(viewsets.ModelViewSet):
    """
    API endpoint para gestionar Usuarios.
    """
    serializer_class = UserSerializer

    def get_queryset(self):
        """
        Sobrescribe el queryset base para añadir filtros.
        """
        queryset = Usuario.objects.all().order_by('id')
        
        # --- Lógica de filtro ---
        email = self.request.query_params.get('email', None)
        if email is not None:
            queryset = queryset.filter(models.Q(username__icontains=email) | models.Q(email__icontains=email))
            
        return queryset

    def get_permissions(self):
        if self.action == 'create':
            permission_classes = [AllowAny]
        elif self.action == 'me' or self.action == 'by_role':
            permission_classes = [IsAuthenticated]
        else:
            permission_classes = [IsAdminUser] 
        return [permission() for permission in permission_classes]

    def create(self, request, *args, **kwargs):
        """
        Maneja el REGISTRO PÚBLICO desde index.html.
        """
        email = request.data.get('email', '').lower().strip()
        password = request.data.get('password')
        rol_name = request.data.get('rol', 'corredor') 

        if not email or not password:
            return Response({"detail": "Email y password son requeridos."}, status=status.HTTP_400_BAD_REQUEST)
        if not email.endswith('@nuam.cl'):
             return Response({"detail": "Solo se permiten correos @nuam.cl"}, status=status.HTTP_400_BAD_REQUEST)
        if Usuario.objects.filter(username=email).exists():
            return Response({"detail": "Este email ya está registrado."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            with transaction.atomic():
                user = Usuario.objects.create_user(
                    username=email, email=email, password=password,
                    first_name=request.data.get('first_name', ''),
                    last_name=request.data.get('last_name', ''),
                    is_active=True 
                )
                group, created = Group.objects.get_or_create(name=rol_name.capitalize())
                user.groups.add(group)
                user.save()
                serializer = UserSerializer(user)
                return Response(serializer.data, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({"detail": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['POST'], permission_classes=[IsAdminUser])
    def admin_create(self, request):
        """
        Crea un usuario desde el panel de admin (dashboard.html), asignando un rol.
        """
        email = request.data.get('email', '').lower().strip()
        rol_name = request.data.get('rol', 'corredor') 
        rut = request.data.get('rut_documento', '').strip()

        if not email or not request.data.get('password'):
            return Response({"detail": "Email y password son requeridos."}, status=status.HTTP_400_BAD_REQUEST)
        
        if Usuario.objects.filter(username=email).exists(): 
            return Response({"detail": "Un usuario con este email (username) ya existe."}, status=status.HTTP_400_BAD_REQUEST)

        if rut and Usuario.objects.filter(rut_documento=rut).exists():
            return Response({"detail": "Un usuario con este RUT/Documento ya existe."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            with transaction.atomic():
                user = Usuario.objects.create_user( 
                    username=email, email=email,
                    password=request.data.get('password'),
                    first_name=request.data.get('first_name', ''),
                    last_name=request.data.get('last_name', ''),
                    genero=request.data.get('genero', ''),
                    telefono=request.data.get('telefono', ''),
                    direccion=request.data.get('direccion', ''),
                    rut_documento=rut if rut else None, 
                    pais=request.data.get('pais', ''),
                    is_active=True 
                )
                if rol_name == 'admin':
                    user.is_staff = True
                    user.is_superuser = True
                else:
                    group, created = Group.objects.get_or_create(name=rol_name.capitalize())
                    user.groups.add(group)
                user.save()
                serializer = CurrentUserSerializer(user)
                return Response(serializer.data, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({"detail": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def update(self, request, *args, **kwargs):
        """
        Maneja la EDICIÓN (PATCH) de un usuario desde el dashboard.
        """
        partial = kwargs.pop('partial', True) 
        instance = self.get_object()
        
        rol_name = request.data.get('rol')
        if rol_name:
            if rol_name == 'admin':
                instance.is_staff = True
                instance.is_superuser = True
                instance.groups.clear() 
            else:
                group, created = Group.objects.get_or_create(name=rol_name.capitalize())
                instance.groups.set([group]) 
                instance.is_staff = False
                instance.is_superuser = False
            instance.save()

        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        if getattr(instance, '_prefetched_objects_cache', None):
            instance._prefetched_objects_cache = {}
        return Response(serializer.data)

    @action(detail=True, methods=['POST'], permission_classes=[IsAdminUser])
    def disable_user(self, request, pk=None):
        try:
            user = self.get_object()
            if user.is_superuser:
                return Response({"detail": "No se puede deshabilitar a un superusuario."}, status=status.HTTP_403_FORBIDDEN)
            user.is_active = False
            user.save()
            return Response({"status": "usuario deshabilitado"}, status=status.HTTP_200_OK)
        except Usuario.DoesNotExist:
            return Response({"detail": "Usuario no encontrado."}, status=status.HTTP_44_NOT_FOUND)

    @action(detail=True, methods=['POST'], permission_classes=[IsAdminUser])
    def enable_user(self, request, pk=None):
        try:
            user = self.get_object()
            user.is_active = True
            user.save()
            return Response({"status": "usuario habilitado"}, status=status.HTTP_200_OK)
        except Usuario.DoesNotExist:
            return Response({"detail": "Usuario no encontrado."}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=True, methods=['DELETE'], permission_classes=[IsAdminUser], name="delete_permanent")
    def delete_permanent(self, request, pk=None):
        try:
            user = self.get_object()
            if user.is_superuser:
                return Response({"detail": "No se puede eliminar a un superusuario."}, status=status.HTTP_403_FORBIDDEN)
            user.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
        except Usuario.DoesNotExist:
            return Response({"detail": "Usuario no encontrado."}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=False, methods=['GET'], permission_classes=[IsAuthenticated])
    def by_role(self, request):
        try:
            admins = Usuario.objects.filter(is_superuser=True)
            supervisors = Usuario.objects.filter(groups__name='Supervisor') 
            corredores = Usuario.objects.filter(groups__name='Corredor')
            
            return Response({
                "administradores": UserSerializer(admins, many=True).data,
                "supervisores": UserSerializer(supervisors, many=True).data,
                "corredores": UserSerializer(corredores, many=True).data,
            })
        except Exception as e:
            return Response({"detail": f"Error al obtener roles. Error: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['GET']) 
    def me(self, request):
        if not request.user.is_authenticated:
            return Response({"detail": "No autenticado"}, status=status.HTTP_401_UNAUTHORIZED)
        serializer = CurrentUserSerializer(request.user)
        return Response(serializer.data)


# =============================================================
# VISTAS DE CALIFICACIONES (CORE)
# =============================================================

class CalificacionViewSet(viewsets.ModelViewSet):
    queryset = Calificacion.objects.all().order_by('-created_at')
    serializer_class = CalificacionSerializer
    permission_classes = [IsOwnerOrAdmin] 
    parser_classes = (MultiPartParser, FormParser, JSONParser,)

    def get_queryset(self):
        user = self.request.user
        queryset = Calificacion.objects.select_related(
            'usuario', 'instrumento', 'mercado', 'estado'
        ).prefetch_related(
            'tributarias', 'tributarias__factores'
        ).filter(is_active=True).order_by('-created_at')

        # Filtros
        calificacion_id = self.request.query_params.get('id', None)
        usuario_email = self.request.query_params.get('usuario_email', None)
        anio = self.request.query_params.get('anio', None)

        if calificacion_id:
            queryset = queryset.filter(id=calificacion_id)
        if usuario_email:
            queryset = queryset.filter(usuario__username__icontains=usuario_email)
        if anio:
            queryset = queryset.filter(fecha_emision__year=anio)

        if user.is_staff or user.groups.filter(name="Supervisor").exists():
            return queryset 
        else:
            return queryset.filter(usuario=user)

    def perform_create(self, serializer):
        serializer.save(usuario=self.request.user, created_by=self.request.user, updated_by=self.request.user)
        logger.info(f"Usuario {self.request.user.username} creó Calificación ID: {serializer.instance.id}")

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)
        logger.info(f"Usuario {self.request.user.username} actualizó Calificación ID: {serializer.instance.id}")

    def perform_destroy(self, instance):
        instance.soft_delete(user=self.request.user)
        logger.info(f"Usuario {self.request.user.username} deshabilitó Calificación ID: {instance.id}")

    # --- 1. ACCIÓN DE EXPORTAR CSV ---
    @action(detail=False, methods=['get'], url_path='exportar_csv')
    def exportar_csv(self, request):
        user = request.user
        queryset = self.get_queryset()

        wb = Workbook()
        ws = wb.active
        ws.title = f"Calificaciones_{user.username}"

        headers = [
            'ID_Calificacion', 'Instrumento_ID', 'Mercado_ID', 'Estado_ID', 'Monto', 
            'Fecha_Emision', 'Fecha_Pago', 'Creado_Por', 'Fecha_Creacion',
            'Secuencia_Trib', 'Evento_Capital', 'Anio_Trib', 'Valor_Historico',
            'Codigo_Factor', 'Valor_Factor'
        ]

        header_font = Font(bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center")
        ws.append(headers) 

        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = header_align
            cell.fill = PatternFill(start_color="FD441E", end_color="FD441E", fill_type="solid")

        for cal in queryset:
            base_data = [
                cal.id, cal.instrumento_id, cal.mercado_id, cal.estado_id, cal.monto_factor,
                cal.fecha_emision, cal.fecha_pago, cal.usuario.username if cal.usuario else 'N/A',
                cal.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            ]
            tributarias = cal.tributarias.all()

            if not tributarias:
                ws.append(base_data + ['', '', '', '', '', '']) 
            else:
                for trib in tributarias:
                    trib_data = [
                        trib.secuencia_evento, trib.evento_capital, trib.anio, trib.valor_historico
                    ]
                    factores = trib.factores.all()
                    if not factores:
                        ws.append(base_data + trib_data + ['', ''])
                    else:
                        for fact in factores:
                            fact_data = [fact.codigo_factor, fact.valor_factor]
                            ws.append(base_data + trib_data + fact_data)

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter 
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="calificaciones_{user.username}.xlsx"'
        wb.save(response) 
        return response

    # --- 2. ACCIÓN DE IMPORTAR CSV (CARGA MASIVA) ---
    @action(detail=False, methods=['POST'])
    def importar_csv(self, request):
        """
        Lee un CSV o XLSX, guarda en BD y genera Log de Auditoría.
        """
        # 1. Validación de Permisos
        es_corredor = request.user.groups.filter(name='Corredor').exists()
        es_admin = request.user.is_staff
        if not es_corredor and not es_admin:
            return Response({"detail": "No tienes permisos."}, status=status.HTTP_403_FORBIDDEN)

        archivo = request.FILES.get('archivo')
        if not archivo:
            return Response({"error": "No se envió ningún archivo."}, status=status.HTTP_400_BAD_REQUEST)

        creados = 0
        errores = []
        rows = []

        try:
            # 2. Leer el archivo (Excel o CSV)
            if archivo.name.endswith('.csv'):
                decoded_file = archivo.read().decode('utf-8').splitlines()
                reader = csv.DictReader(decoded_file)
                rows = list(reader)

            elif archivo.name.endswith('.xlsx'):
                wb = load_workbook(filename=archivo, read_only=True)
                ws = wb.active
                # Leer encabezados de la fila 1
                headers = [cell.value for cell in ws[1]]
                # Leer filas desde la 2
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_data = dict(zip(headers, row))
                    rows.append(row_data)
            else:
                return Response({'error': 'Formato no soportado. Use .csv o .xlsx'}, status=status.HTTP_400_BAD_REQUEST)

            # 3. Procesar filas y guardar en BD
            with transaction.atomic():
                for i, row in enumerate(rows):
                    try:
                        # Mapeo flexible de columnas
                        inst_id = row.get('Instrumento_ID') or row.get('instrumento_id')
                        merc_id = row.get('Mercado_ID') or row.get('mercado_id')
                        est_id = row.get('Estado_ID') or row.get('estado_id')
                        monto = row.get('Monto') or row.get('monto_factor')
                        f_emision = row.get('Fecha_Emision') or row.get('fecha_emision')
                        f_pago = row.get('Fecha_Pago') or row.get('fecha_pago')

                        if not inst_id: continue 

                        Calificacion.objects.create(
                            usuario=request.user,
                            instrumento_id=inst_id,
                            mercado_id=merc_id,
                            estado_id=est_id,
                            monto_factor=monto,
                            fecha_emision=f_emision,
                            fecha_pago=f_pago,
                            created_by=request.user,
                            updated_by=request.user
                        )
                        creados += 1
                    except Exception as e:
                        errores.append(f"Fila {i + 2}: {str(e)}")

                # 4. CREAR LOG DE HISTORIAL
                if creados > 0:
                    Log.objects.create(
                        usuario=request.user,
                        accion="Carga Masiva",
                        detalle=f"Archivo '{archivo.name}' procesado. {creados} registros creados."
                    )

            # 5. Respuesta final
            return Response({
                "creados": creados,
                "errores": errores,
                "mensaje": "Carga procesada correctamente"
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
    @action(detail=False, methods=['GET'], permission_classes=[IsOwnerOrAdmin])
    def mis_calificaciones(self, request):
        mis_calificaciones = self.get_queryset().filter(usuario=request.user, is_active=True)
        serializer = self.get_serializer(mis_calificaciones, many=True)
        return Response(serializer.data)

# =============================================================
# VISTAS DE MODELOS GENÉRICOS
# =============================================================

class InstrumentoViewSet(viewsets.ModelViewSet):
    queryset = Instrumento.objects.all()
    serializer_class = InstrumentoSerializer
    permission_classes = [IsAdminOrReadOnly]

class MercadoViewSet(viewsets.ModelViewSet):
    queryset = Mercado.objects.all()
    serializer_class = MercadoSerializer
    permission_classes = [IsAdminOrReadOnly]

class EstadoViewSet(viewsets.ModelViewSet):
    queryset = Estado.objects.all()
    serializer_class = EstadoSerializer
    permission_classes = [IsAdminOrReadOnly]

class ArchivoViewSet(viewsets.ModelViewSet):
    queryset = Archivo.objects.all()
    serializer_class = ArchivoSerializer
    permission_classes = [IsAdminOrReadOnly]

class CalificacionTributariaViewSet(viewsets.ModelViewSet):
    queryset = CalificacionTributaria.objects.all()
    serializer_class = CalificacionTributariaSerializer
    permission_classes = [IsOwnerOrAdmin] 

class FactorTributarioViewSet(viewsets.ModelViewSet):
    queryset = FactorTributario.objects.all()
    serializer_class = FactorTributarioSerializer
    permission_classes = [IsOwnerOrAdmin]

class LogViewSet(viewsets.ReadOnlyModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = LogSerializer
    
    def get_queryset(self):
        user = self.request.user
        if user.is_staff or user.groups.filter(name="Supervisor").exists():
            return Log.objects.all().order_by('-fecha')
        else:
            return Log.objects.filter(usuario=user).order_by('-fecha')

class AuditoriaViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = AuditoriaSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.is_staff:
            return Auditoria.objects.all().order_by('-fecha')
        else:
            return Auditoria.objects.filter(usuario=user).order_by('-fecha')

class RespaldoViewSet(viewsets.ModelViewSet):
    """
    API endpoint para gestionar los registros de Respaldos.
    """
    queryset = Respaldo.objects.all()
    serializer_class = RespaldoSerializer
    permission_classes = [IsAdminUser]